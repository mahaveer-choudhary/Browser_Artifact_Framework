import os
import json
import threading
import webbrowser
from flask import Flask, render_template, jsonify, request, send_from_directory, send_file

app = Flask(__name__, 
            static_folder='dashboard',
            template_folder='dashboard')

PORT = 5000

import sys

if getattr(sys, 'frozen', False):
    DATA_DIR = sys._MEIPASS
    EXE_DIR = os.path.dirname(sys.executable)
else:
    DATA_DIR = os.path.dirname(os.path.abspath(__file__))
    EXE_DIR = DATA_DIR

DUMPS_DIR = os.path.join(EXE_DIR, 'dumps')

@app.route('/<path:filename>')
def serve_dashboard_files(filename):
    return send_from_directory(os.path.join(DATA_DIR, 'dashboard'), filename)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/files', methods=['GET'])
def get_files():
    """List available JSON files."""
    files = []
    dumps_path = DUMPS_DIR
    if os.path.exists(dumps_path):
        for f in os.listdir(dumps_path):
            if f.endswith('.json') and 'secrets' in f:
                files.append(f)
    return jsonify(files)

@app.route('/api/data/<filename>', methods=['GET'])
def get_data(filename):
    """Get data from a specific JSON file."""
    if not filename.endswith('.json'):
        return jsonify({"error": "Invalid file type"}), 400
        
    path = os.path.join(DATA_DIR, 'dumps', filename)
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return jsonify(data)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    return jsonify({"error": "File not found"}), 404

    return jsonify({"error": "File not found"}), 404

@app.route('/reports/<path:filename>')
def serve_report(filename):
    """Serve generated reports from dumps folder."""
    return send_from_directory(DUMPS_DIR, filename)

@app.route('/api/extract/<browser>', methods=['POST'])
def trigger_extraction(browser):
    """Trigger extraction for a specific browser."""
    if not os.path.exists(DUMPS_DIR):
        os.makedirs(DUMPS_DIR)
    
    output_file = os.path.join(DUMPS_DIR, f"secrets_{browser}.json")
    
    try:
        import subprocess
        
        # Kill browser processes first
        subprocess.run(
            "taskkill /F /IM chrome.exe /IM msedge.exe /IM brave.exe /IM firefox.exe", 
            shell=True, 
            stdout=subprocess.DEVNULL, 
            stderr=subprocess.DEVNULL
        )
        
        # Run extraction natively instead of spawning a new Python subprocess
        from browsers.chromium import ChromiumExtractor
        from browsers.firefox import FirefoxExtractor
        from utils.forensics import log_chain_of_custody
        
        b = browser.lower()
        full_data = {}
        log_output = []
        
        if b == "firefox":
            extractor = FirefoxExtractor()
            logins = extractor.extract_logins()
            if logins:
                cookies = extractor.extract_cookies()
                history = extractor.extract_history()
                extensions = extractor.extract_extensions()
                
                from utils.intelligence import BAFIntelligence
                intel = BAFIntelligence(b)
                cloud_tokens = intel.identify_cloud_tokens(cookies)
                iocs = intel.scan_iocs(history)
                
                full_data[b] = {
                    "logins": logins,
                    "cookies": cookies,
                    "history": history,
                    "intelligence": {
                        "extensions": extensions,
                        "cloud_tokens": cloud_tokens,
                        "iocs": iocs,
                        "deleted_history": []
                    }
                }
                log_output.append(f"[+] Extracted {len(logins)} logins, {len(cookies)} cookies, {len(history)} history.")
            else:
                log_output.append("[-] No Firefox data found or decryption failed.")
        else:
            extractor = ChromiumExtractor(b)
            if extractor.get_keys():
                cookies = extractor.extract_cookies()
                logins = extractor.extract_logins()
                history = extractor.extract_history()
                web_data = extractor.extract_web_data()
                bookmarks = extractor.extract_bookmarks()
                
                from utils.intelligence import BAFIntelligence
                intel = BAFIntelligence(b)
                extensions = intel.analyze_extensions()
                cloud_tokens = intel.identify_cloud_tokens(cookies)
                iocs = intel.scan_iocs(history)
                deleted = getattr(extractor, 'deleted_history', [])
                
                full_data[b] = {
                    "cookies": cookies,
                    "logins": logins,
                    "history": history,
                    "credit_cards": web_data.get("credit_cards", []),
                    "autofill": web_data.get("autofill", []),
                    "bookmarks": bookmarks,
                    "intelligence": {
                        "extensions": extensions,
                        "cloud_tokens": cloud_tokens,
                        "iocs": iocs,
                        "deleted_history": deleted
                    }
                }
                log_output.append(f"[+] Extracted {len(cookies)} cookies, {len(logins)} logins, {len(history)} history items.")
            else:
                log_output.append("[-] Could not retrieve keys. Skipping.")
                if b in ["chrome", "edge", "brave"]:
                    log_output.append("App-Bound key missing. Background browser processes might be interfering.")
                
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(full_data, f, indent=4)
            
        file_hash = log_chain_of_custody(output_file, DUMPS_DIR)
        
        from utils.intelligence import BAFIntelligence
        timeline_path = BAFIntelligence.generate_super_timeline(full_data, DUMPS_DIR)
        if timeline_path:
            log_output.append(f"[+] Super Timeline CSV generated.")
            
        system_file = os.path.join(DUMPS_DIR, "system_artifacts.json")
        if not os.path.exists(system_file):
            from utils.system_artifacts import SystemArtifacts
            sys_data = SystemArtifacts.run_all()
            with open(system_file, "w", encoding="utf-8") as f:
                json.dump(sys_data, f, indent=4)
            log_output.append("[+] System Artifacts extracted (sys_artifacts.json).")
            
        log_output.append(f"[+] Data saved to {os.path.basename(output_file)}")
        log_output.append(f"[+] SHA-256: {file_hash}")
        
        return jsonify({
            "status": "success", 
            "file": os.path.basename(output_file),
            "log": "\n".join(log_output)
        })
        
    except Exception as e:
        return jsonify({
            "status": "error", 
            "log": str(e)
        })

@app.route('/api/system_artifacts', methods=['GET'])
def get_system_artifacts():
    """Retrieve system artifacts from the JSON dump."""
    system_file = os.path.join(DUMPS_DIR, "system_artifacts.json")
    if os.path.exists(system_file):
        try:
            with open(system_file, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        except Exception:
            return jsonify({})
    return jsonify({})

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get combined stats from all browsers."""
    stats = {
        "total_logins": 0,
        "total_cookies": 0,
        "total_history": 0,
        "browsers": []
    }
    
    browsers = ['chrome', 'edge', 'firefox', 'brave', 'opera', 'vivaldi', 'duckduckgo']
    
    for browser in browsers:
        path = os.path.join(DATA_DIR, 'dumps', f"secrets_{browser}.json")
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    browser_data = data.get(browser, {})
                    
                    logins = len(browser_data.get('logins', []))
                    cookies = len(browser_data.get('cookies', []))
                    history = len(browser_data.get('history', []))
                    
                    stats["total_logins"] += logins
                    stats["total_cookies"] += cookies
                    stats["total_history"] += history
                    stats["browsers"].append({
                        "name": browser,
                        "logins": logins,
                        "cookies": cookies,
                        "history": history
                    })
            except Exception:
                pass
    
    return jsonify(stats)

@app.route('/api/export', methods=['POST'])
def export_report():
    """Generate a forensic PDF, CSV, or ZIP report."""
    try:
        options = request.get_json() or {}
        format_type = options.get('format', 'pdf')
        selected_browsers = options.get('browsers', ['chrome', 'edge', 'firefox', 'brave', 'opera', 'vivaldi'])
        
        inc_logins = options.get('logins', True)
        inc_cookies = options.get('cookies', True)
        inc_history = options.get('history', True)
        inc_system = options.get('system', True)
        
        full_data = {}
        for browser in selected_browsers:
            path = os.path.join(DUMPS_DIR, f"secrets_{browser}.json")
            if os.path.exists(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        b_data = data.get(browser, {})
                        
                        filtered_b = {}
                        if inc_logins: filtered_b['logins'] = b_data.get('logins', [])
                        if inc_cookies: filtered_b['cookies'] = b_data.get('cookies', [])
                        if inc_history: filtered_b['history'] = b_data.get('history', [])
                        if 'intelligence' in b_data: filtered_b['intelligence'] = b_data['intelligence']
                        
                        full_data[browser] = filtered_b
                except: pass
                
        system_data = {}
        if inc_system:
            system_file = os.path.join(DUMPS_DIR, "system_artifacts.json")
            if os.path.exists(system_file):
                with open(system_file, 'r', encoding='utf-8') as f:
                    system_data = json.load(f)
                    
        import zipfile
        import csv
        
        import openpyxl
        import re
        
        def clean_xml_string(s):
            """Remove control characters invalid in XML to prevent Excel corruption."""
            if not s: return ""
            return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', str(s))
            
        # 1. Generate Dynamic Multi-Sheet Excel Workbook
        excel_path = os.path.join(DUMPS_DIR, "Forensic_Report.xlsx")
        wb = openpyxl.Workbook()
        
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        for b, d in full_data.items():
            if not d.get('logins') and not d.get('history'): continue
            # Excel limits sheet titles to 31 chars
            safe_title = b.upper()[:30]
            ws = wb.create_sheet(title=safe_title)
            ws.append(["Type", "Title / Hostname", "Value / Field"])
            
            if inc_logins:
                for l in d.get('logins', []):
                    ws.append([
                        "Login", 
                        clean_xml_string(l.get('origin') or l.get('action_url') or l.get('hostname') or 'N/A'), 
                        clean_xml_string(l.get('username', ''))
                    ])
            if inc_history:
                for h in d.get('history', []):
                    ws.append([
                        "History", 
                        clean_xml_string(h.get('title', 'No Title')), 
                        clean_xml_string(h.get('url', ''))
                    ])
                    
        if inc_system and system_data:
            ws_sys = wb.create_sheet(title="SYSTEM_OS")
            ws_sys.append(["Category", "Item Description / SSID", "Path / Password"])
            for w in system_data.get('wifi_profiles', []):
                ws_sys.append(["Wi-Fi Profile", clean_xml_string(w.get('ssid', '')), clean_xml_string(w.get('password', ''))])
            for a in system_data.get('anti_forensics', []):
                ws_sys.append(["Anti-Forensics / Tool", clean_xml_string(a.get('name', '')), clean_xml_string(a.get('path', ''))])
            for u in system_data.get('usb_history', []):
                ws_sys.append(["USB Device", clean_xml_string(u.get('name', '')), clean_xml_string(u.get('serial_number', ''))])
                
        if not wb.sheetnames:
            wb.create_sheet(title="Empty_Report").append(["No data selected for export"])
            
        wb.save(excel_path)

        # 2. Build output based on format
        if format_type == 'pdf':
            from utils.reporting import generate_pdf_report
            pdf_path = os.path.join(DUMPS_DIR, "Forensic_Report.pdf")
            generate_pdf_report(full_data, system_data, pdf_path)
            return send_file(pdf_path, as_attachment=True, download_name="BAF_Forensic_Report.pdf")
            
        elif format_type == 'xlsx':
            return send_file(excel_path, as_attachment=True, download_name="BAF_Forensic_Export.xlsx")
            
        else: # ZIP
            from utils.reporting import generate_pdf_report
            pdf_path = os.path.join(DUMPS_DIR, "Forensic_Report.pdf")
            generate_pdf_report(full_data, system_data, pdf_path)
            
            zip_path = os.path.join(DUMPS_DIR, "Full_Forensic_Bundle.zip")
            with zipfile.ZipFile(zip_path, 'w') as zf:
                if os.path.exists(pdf_path): zf.write(pdf_path, arcname="Forensic_Report.pdf")
                if os.path.exists(excel_path): zf.write(excel_path, arcname="Super_Timeline.xlsx")
                for f in os.listdir(DUMPS_DIR):
                    if f.endswith('.json') and 'secrets' in f:
                        zf.write(os.path.join(DUMPS_DIR, f), arcname=f"raw_data/{f}")
            return send_file(zip_path, as_attachment=True, download_name="BAF_Full_Forensic_Bundle.zip")
        
    except Exception as e:
        return jsonify({"status": "error", "log": str(e)})


@app.route('/api/search', methods=['POST'])
def search_artifacts():
    """Global eDiscovery search across all extracted browser data."""
    try:
        data = request.get_json() or {}
        keyword = (data.get('keyword', '')).lower()
        if not keyword:
            return jsonify({"status": "error", "message": "No keyword provided"})
            
        results = {
            "logins": [],
            "cookies": [],
            "history": []
        }
        
        dumps_path = DUMPS_DIR
        if os.path.exists(dumps_path):
            for f in os.listdir(dumps_path):
                if f.endswith('.json') and f.startswith('secrets_'):
                    browser = f.replace('secrets_', '').replace('.json', '')
                    path = os.path.join(dumps_path, f)
                    try:
                        with open(path, 'r', encoding='utf-8') as file:
                            jdata = json.load(file)
                            bdata = jdata.get(browser, {})
                            
                            # Logins
                            for item in bdata.get('logins', []):
                                if keyword in str(item.get('origin', '')).lower() or keyword in str(item.get('username', '')).lower() or keyword in str(item.get('password', '')).lower():
                                    item['browserSource'] = browser
                                    results['logins'].append(item)
                                    
                            # Cookies
                            for item in bdata.get('cookies', []):
                                if keyword in str(item.get('host', '')).lower() or keyword in str(item.get('name', '')).lower() or keyword in str(item.get('value', '')).lower():
                                    item['browserSource'] = browser
                                    results['cookies'].append(item)
                                    
                            # History
                            for item in bdata.get('history', []):
                                if keyword in str(item.get('url', '')).lower() or keyword in str(item.get('title', '')).lower():
                                    item['browserSource'] = browser
                                    results['history'].append(item)
                    except Exception:
                        pass
                        
        return jsonify({"status": "success", "results": results})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})


def open_browser():
    """Open browser after a short delay."""
    import time
    time.sleep(1)
    webbrowser.open(f'http://127.0.0.1:{PORT}')

def run_BAF(open_browser_flag=True):
    """Run the dashboard server."""
    print(f"\n{'='*50}")
    print(f"  Browser Artifact Framework (BAF) Dashboard")
    print(f"  Running on: http://127.0.0.1:{PORT}")
    print(f"{'='*50}\n")
    
    if open_browser_flag:
        threading.Thread(target=open_browser, daemon=True).start()
    
    app.run(
        host='127.0.0.1', 
        port=PORT, 
        debug=False,
        threaded=True
    )

if __name__ == '__main__':
    run_BAF()
